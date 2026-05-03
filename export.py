"""Phase 4: Export items and lots to Excel with thumbnails, hyperlinks, and eBay File Exchange sheet."""

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

import db
from config import OUTPUT_DIR

MAIN_HEADERS = [
    "item_id", "status", "thumbnail",
    "item_type", "probable_maker", "makers_marks_observed", "probable_era",
    "material", "style_keywords", "condition_notes",
    "confidence_score", "confidence_reasoning", "suggested_search_query", "needs_review",
    "comp_low", "comp_median", "comp_high", "comp_sample_size", "comp_confidence",
    "comp_links",
    "ebay_title", "etsy_title", "fb_marketplace_title",
    "description", "suggested_price", "suggested_platform", "tags", "category_ebay",
]

LOT_HEADERS = [
    "lot_id", "lot_name", "thumbnail", "item_ids", "lot_price",
    "suggested_platform", "status", "lot_description",
]

EBAY_FX_HEADERS = [
    "Action", "ItemID", "Category", "Title", "Description",
    "StartPrice", "BuyItNowPrice", "Quantity", "ConditionID", "PicURL",
]

HEADER_FILL    = PatternFill("solid", fgColor="2D2D2D")
HEADER_FONT    = Font(color="FFFFFF", bold=True)
LOT_FILL       = PatternFill("solid", fgColor="1A3A5C")
YARD_FILL      = PatternFill("solid", fgColor="4A2020")
REVIEWED_FILL  = PatternFill("solid", fgColor="1A3A1A")


def _make_thumbnail(photo_path: str, size=(80, 80)) -> Path | None:
    try:
        src = Path(photo_path)
        if not src.exists():
            from config import PROCESSED_PHOTOS_DIR
            alt = PROCESSED_PHOTOS_DIR / src.name
            src = alt if alt.exists() else src
        if not src.exists():
            return None
        thumb_dir = OUTPUT_DIR / "thumbs"
        thumb_dir.mkdir(exist_ok=True)
        dst = thumb_dir / f"{src.stem}_thumb.png"
        with PILImage.open(src) as img:
            img.thumbnail(size, PILImage.LANCZOS)
            img.save(dst, "PNG")
        return dst
    except Exception:
        return None


def _write_header_row(ws, headers: list[str], fill, font):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20


def _embed_image(ws, path: Path | None, col_letter: str, row: int, size=(60, 60)):
    if path and Path(path).exists():
        try:
            img = XLImage(str(path))
            img.width, img.height = size
            ws.add_image(img, f"{col_letter}{row}")
        except Exception:
            pass


def export_to_excel(items: list[dict] | None = None) -> Path:
    db.init_db()
    all_items = items if items is not None else db.get_all_items()
    all_lots  = db.get_all_lots()

    # Partition items: lotted → skip individual export | yard_sale → own section | normal → main sheet
    individual_items = [i for i in all_items if not i.get("lot_id") and not i.get("yard_sale")]
    yard_sale_items  = [i for i in all_items if i.get("yard_sale")]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Individual Items ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Individual Items"
    _write_header_row(ws, MAIN_HEADERS, HEADER_FILL, HEADER_FONT)

    for row_idx, item in enumerate(individual_items, 2):
        ws.row_dimensions[row_idx].height = 65
        photo_paths  = json.loads(item.get("photo_paths") or "[]")
        thumb_path   = _make_thumbnail(photo_paths[0]) if photo_paths else None
        comp_listings = json.loads(item.get("comp_listings") or "[]")
        comp_links   = " | ".join(l.get("url", "") for l in comp_listings[:5] if l.get("url"))

        row_data = {
            "item_id": item.get("item_id"),
            "status": item.get("status"),
            "thumbnail": "",
            "item_type": item.get("item_type"),
            "probable_maker": item.get("probable_maker"),
            "makers_marks_observed": item.get("makers_marks_observed"),
            "probable_era": item.get("probable_era"),
            "material": item.get("material"),
            "style_keywords": ", ".join(json.loads(item.get("style_keywords") or "[]")),
            "condition_notes": item.get("condition_notes"),
            "confidence_score": item.get("confidence_score"),
            "confidence_reasoning": item.get("confidence_reasoning"),
            "suggested_search_query": item.get("suggested_search_query"),
            "needs_review": "YES" if item.get("needs_review") else "",
            "comp_low": item.get("comp_low"),
            "comp_median": item.get("comp_median"),
            "comp_high": item.get("comp_high"),
            "comp_sample_size": item.get("comp_sample_size"),
            "comp_confidence": item.get("comp_confidence"),
            "comp_links": comp_links,
            "ebay_title": item.get("ebay_title"),
            "etsy_title": item.get("etsy_title"),
            "fb_marketplace_title": item.get("fb_marketplace_title"),
            "description": item.get("description"),
            "suggested_price": item.get("suggested_price"),
            "suggested_platform": item.get("suggested_platform"),
            "tags": ", ".join(json.loads(item.get("tags") or "[]")),
            "category_ebay": item.get("category_ebay"),
        }

        row_fill = REVIEWED_FILL if item.get("status") == "reviewed" else None
        thumb_col = get_column_letter(MAIN_HEADERS.index("thumbnail") + 1)

        for col_idx, header in enumerate(MAIN_HEADERS, 1):
            if header == "thumbnail":
                _embed_image(ws, thumb_path, thumb_col, row_idx)
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_fill:
                    cell.fill = row_fill

    for col_idx in range(1, len(MAIN_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.column_dimensions["C"].width = 10
    desc_col = get_column_letter(MAIN_HEADERS.index("description") + 1)
    ws.column_dimensions[desc_col].width = 50

    # ── Sheet 2: Lots ─────────────────────────────────────────────────────────
    ws_lots = wb.create_sheet("Lots")
    _write_header_row(ws_lots, LOT_HEADERS, LOT_FILL, Font(color="FFFFFF", bold=True))

    for row_idx, lot in enumerate(all_lots, 2):
        ws_lots.row_dimensions[row_idx].height = 80
        item_ids = json.loads(lot.get("item_ids") or "[]")
        thumb_col = get_column_letter(LOT_HEADERS.index("thumbnail") + 1)
        thumb_path = Path(lot.get("thumbnail_path") or "")

        row_data = {
            "lot_id": lot.get("lot_id"),
            "lot_name": lot.get("lot_name"),
            "thumbnail": "",
            "item_ids": ", ".join(item_ids),
            "lot_price": lot.get("lot_price"),
            "suggested_platform": lot.get("suggested_platform"),
            "status": lot.get("status"),
            "lot_description": lot.get("lot_description"),
        }
        for col_idx, header in enumerate(LOT_HEADERS, 1):
            if header == "thumbnail":
                _embed_image(ws_lots, thumb_path if thumb_path.exists() else None,
                             thumb_col, row_idx, size=(70, 70))
            else:
                cell = ws_lots.cell(row=row_idx, column=col_idx, value=row_data.get(header))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx in range(1, len(LOT_HEADERS) + 1):
        ws_lots.column_dimensions[get_column_letter(col_idx)].width = 22
    ws_lots.column_dimensions[get_column_letter(LOT_HEADERS.index("lot_description") + 1)].width = 55

    # ── Sheet 3: Yard Sale ────────────────────────────────────────────────────
    if yard_sale_items:
        ws_yard = wb.create_sheet("Yard Sale (FB)")
        yard_headers = ["item_id", "item_type", "probable_maker", "probable_era",
                        "condition_notes", "suggested_price", "fb_marketplace_title", "description"]
        _write_header_row(ws_yard, yard_headers, YARD_FILL, Font(color="FFFFFF", bold=True))
        for row_idx, item in enumerate(yard_sale_items, 2):
            for col_idx, header in enumerate(yard_headers, 1):
                cell = ws_yard.cell(row=row_idx, column=col_idx, value=item.get(header))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx in range(1, len(yard_headers) + 1):
            ws_yard.column_dimensions[get_column_letter(col_idx)].width = 22

    # ── Sheet 4: eBay File Exchange ───────────────────────────────────────────
    ws_fx = wb.create_sheet("eBay File Exchange")
    for col_idx, h in enumerate(EBAY_FX_HEADERS, 1):
        ws_fx.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

    fx_row = 2
    # Individual items
    for item in individual_items:
        photo_paths = json.loads(item.get("photo_paths") or "[]")
        ws_fx.cell(fx_row, 1, "Add")
        ws_fx.cell(fx_row, 3, item.get("category_ebay"))
        ws_fx.cell(fx_row, 4, item.get("ebay_title"))
        ws_fx.cell(fx_row, 5, item.get("description"))
        ws_fx.cell(fx_row, 6, item.get("suggested_price"))
        ws_fx.cell(fx_row, 7, item.get("suggested_price"))
        ws_fx.cell(fx_row, 8, 1)
        ws_fx.cell(fx_row, 10, photo_paths[0] if photo_paths else "")
        fx_row += 1
    # Lots
    for lot in all_lots:
        ws_fx.cell(fx_row, 1, "Add")
        ws_fx.cell(fx_row, 4, lot.get("lot_name"))
        ws_fx.cell(fx_row, 5, lot.get("lot_description"))
        ws_fx.cell(fx_row, 6, lot.get("lot_price"))
        ws_fx.cell(fx_row, 7, lot.get("lot_price"))
        ws_fx.cell(fx_row, 8, 1)
        ws_fx.cell(fx_row, 10, lot.get("thumbnail_path", ""))
        fx_row += 1

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"antique_pipeline_{ts}.xlsx"
    wb.save(out_path)
    return out_path
