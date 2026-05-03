"""
Master Listings Spreadsheet Generator.
Run: python master_export.py
Output: output/master_listings.xlsx
"""

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env", override=True)

import db
from config import OUTPUT_DIR, PROCESSED_PHOTOS_DIR

OUTPUT_FILE = OUTPUT_DIR / "master_listings.xlsx"

# ── Palette ───────────────────────────────────────────────────────────────────
class C:
    EBAY_HDR    = "2F75B6"; EBAY_LIGHT   = "DDEEFF"
    ETSY_HDR    = "C55A11"; ETSY_LIGHT   = "FFF0E8"
    FB_HDR      = "1F3864"; FB_LIGHT     = "E8F0FE"
    SPEC_HDR    = "7030A0"; SPEC_LIGHT   = "F3E6FF"
    CROSS_HDR   = "375623"; CROSS_LIGHT  = "EBF5E0"
    REVIEW_HDR  = "C00000"; REVIEW_LIGHT = "FFE0E0"
    WHITE       = "FFFFFF"; GRAY         = "F2F2F2"
    SEC_A = "BDD7EE"; SEC_B = "E2EFDA"; SEC_C = "FCE4D6"
    SEC_D = "EAD1DC"; SEC_E = "F2F2F2"
    LOW_CONF    = "FFF2CC"  # yellow row tint for confidence < 70
    FLAG_ROW    = "FFE0CC"  # orange row tint for special notes

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size)
def _align(wrap=True, h="left", v="top"): return Alignment(wrap_text=wrap, horizontal=h, vertical=v)


# ── Column definitions ─────────────────────────────────────────────────────────
EBAY_COLS = [
    # (header_label, field_key, section, col_width)
    # Section A — Listing Basics
    ("Row ID",              "row_id",           "A", 8),
    ("Listing Title",       "ebay_title",        "A", 55),
    ("Subtitle",            "subtitle",          "A", 40),
    ("Full HTML Description","html_description", "A", 70),
    ("Price (USD)",         "price",             "A", 10),
    ("Format",              "format",            "A", 12),
    ("Duration",            "duration",          "A", 10),
    ("Quantity",            "quantity",          "A", 8),
    ("Condition",           "condition",         "A", 18),
    ("Condition Description","condition_desc",   "A", 40),
    # Section B — Item Specifics
    ("UPC",                 "upc",               "B", 18),
    ("Material",            "material",          "B", 22),
    ("Subject",             "subject",           "B", 18),
    ("Theme",               "theme",             "B", 22),
    ("Color",               "color",             "B", 14),
    ("Origin",              "origin",            "B", 20),
    ("Type",                "item_type_short",   "B", 22),
    ("Production Style",    "production_style",  "B", 22),
    ("Production Technique","production_tech",   "B", 22),
    ("Backstamp/Marks",     "marks",             "B", 28),
    ("Country of Origin",   "country_origin",    "B", 18),
    ("Features",            "features",          "B", 28),
    ("Time Period",         "time_period",       "B", 20),
    ("MPN",                 "mpn",               "B", 14),
    ("Pattern",             "pattern",           "B", 16),
    ("Product Line",        "product_line",      "B", 22),
    ("Year Manufactured",   "year_mfr",          "B", 16),
    ("Style",               "style",             "B", 22),
    ("Finish",              "finish",            "B", 18),
    # Section C — Shipping
    ("Est. Weight (oz)",    "weight_oz",         "C", 14),
    ("Suggested Box Size",  "box_size",          "C", 30),
    ("Shipping Service",    "ship_svc",          "C", 24),
    ("Fragile",             "fragile",           "C", 8),
    ("Packing Notes",       "pack_notes",        "C", 38),
    ("Shipping Type",       "ship_type",         "C", 14),
    ("Handling Time",       "handling",          "C", 14),
    ("Location",            "location",          "C", 18),
    # Section D — Photos
    ("Photo 1",             "photo_1",           "D", 50),
    ("Photo 2",             "photo_2",           "D", 50),
    ("Photo 3",             "photo_3",           "D", 50),
    ("Photo 4",             "photo_4",           "D", 50),
    ("Photo 5",             "photo_5",           "D", 50),
    ("Photo Order Notes",   "photo_notes",       "D", 38),
    # Section E — Internal
    ("Confidence Score",    "confidence",        "E", 14),
    ("Comp Low",            "comp_low",          "E", 10),
    ("Comp Median",         "comp_median",       "E", 12),
    ("Comp High",           "comp_high",         "E", 10),
    ("Comp Sample Size",    "comp_n",            "E", 14),
    ("Secondary Platforms", "secondary",         "E", 28),
    ("Lot ID",              "lot_id",            "E", 18),
    ("Special Notes",       "special_notes",     "E", 50),
]

ETSY_COLS = [
    ("Row ID",              "row_id",            12),
    ("Listing Title",       "etsy_title",        70),
    ("Description",         "etsy_description",  80),
    ("Price (USD)",         "price",             10),
    ("Tags",                "tags",              60),
    ("Materials",           "materials",         30),
    ("Category Path",       "etsy_category",     40),
    ("When Was It Made",    "etsy_era",          20),
    ("Who Made It",         "who_made_it",       28),
    ("Quantity",            "quantity",          10),
    ("Renewal Option",      "renewal",           14),
    ("Shipping Profile",    "ship_profile",      32),
    ("Shop Section",        "etsy_section",      24),
    ("Photo 1",             "photo_1",           50),
    ("Photo 2",             "photo_2",           50),
    ("Photo 3",             "photo_3",           50),
    ("Photo 4",             "photo_4",           50),
    ("Photo 5",             "photo_5",           50),
    ("Photo Order Notes",   "photo_notes",       38),
    ("Secondary Platforms", "secondary",         28),
    ("Special Notes",       "special_notes",     50),
]

FB_COLS = [
    ("Row ID",              "row_id",            10),
    ("Title",               "fb_title",          45),
    ("Description",         "fb_description",    70),
    ("Price (USD)",         "fb_price",          10),
    ("Category",            "fb_category",       28),
    ("Condition",           "condition",         18),
    ("Photo 1",             "photo_1",           50),
    ("Photo 2",             "photo_2",           50),
    ("Pickup Location",     "location",          22),
    ("Negotiation Floor",   "neg_floor",         16),
    ("Special Notes",       "special_notes",     40),
]

SPEC_COLS = [
    ("Row ID",              "row_id",            10),
    ("Platform",            "spec_platform",     16),
    ("Title",               "spec_title",        55),
    ("Description",         "spec_description",  80),
    ("Price (USD)",         "spec_price",        10),
    ("Category",            "spec_category",     32),
    ("Key Selling Points",  "key_points",        50),
    ("Photo 1",             "photo_1",           50),
    ("Photo 2",             "photo_2",           50),
    ("Special Requirements","spec_requirements", 45),
]

CROSS_COLS = [
    ("Item / Lot Name",     "name",              40),
    ("Platform 1",          "platform_1",        16),
    ("Sheet Row (P1)",      "row_p1",            14),
    ("Platform 2",          "platform_2",        16),
    ("Sheet Row (P2)",      "row_p2",            14),
    ("Platform 3",          "platform_3",        16),
    ("Sheet Row (P3)",      "row_p3",            14),
    ("Differentiation Notes","diff_notes",       80),
]

REVIEW_COLS = [
    ("Row ID",              "row_id",            10),
    ("Item ID",             "item_id",           32),
    ("Item Type",           "item_type",         32),
    ("Probable Maker",      "maker",             24),
    ("Confidence Score",    "confidence",        14),
    ("Issue",               "issue",             40),
    ("Suggested Price",     "price",             12),
    ("Photos",              "photo_1",           50),
    ("Special Notes",       "special_notes",     50),
]

SEC_COLORS = {"A": C.SEC_A, "B": C.SEC_B, "C": C.SEC_C, "D": C.SEC_D, "E": C.SEC_E}


# ── Helper functions ──────────────────────────────────────────────────────────

def resolve_photo(raw: str) -> str:
    if not raw: return ""
    p = Path(raw)
    if p.exists(): return str(p)
    alt = PROCESSED_PHOTOS_DIR / p.name
    return str(alt) if alt.exists() else str(p)


def get_photos(row: dict, is_lot: bool) -> list[str]:
    if is_lot:
        thumb = resolve_photo(row.get("thumbnail_path") or "")
        item_ids = json.loads(row.get("item_ids") or "[]")
        extras = []
        for iid in item_ids[:4]:
            item = db.get_item(iid)
            if item:
                paths = json.loads(item.get("photo_paths") or "[]")
                if paths:
                    extras.append(resolve_photo(paths[0]))
        return ([thumb] + extras)[:5] if thumb else extras[:5]
    else:
        paths = json.loads(row.get("photo_paths") or "[]")
        return [resolve_photo(p) for p in paths[:5]]


def photo_dict(photos: list[str]) -> dict:
    d = {}
    for i in range(1, 6):
        d[f"photo_{i}"] = photos[i-1] if i <= len(photos) else ""
    return d


def fmt_html(raw: str, lot_items: list[str] = None) -> str:
    if not raw: return ""
    text = raw.replace("\r\n", "\n").replace("\r", "\n")
    paras = [p.strip() for p in re.split(r"\n{2,}", text.strip()) if p.strip()]
    html = ["".join(f"<p>{p.replace(chr(10), '<br>')}</p>" for p in paras)]
    if lot_items:
        html.append("<hr><p><strong>This lot includes:</strong></p><ul>")
        html.extend(f"<li>{li}</li>" for li in lot_items)
        html.append("</ul><p><em>Sold as a set only — will not separate.</em></p>")
    return "".join(html)


def trunc(text: str, n: int) -> str:
    if not text: return ""
    t = text.strip()
    return t if len(t) <= n else t[:n].rsplit(" ", 1)[0]


def weight_oz(material: str, item_type: str) -> int:
    m, t = (material or "").lower(), (item_type or "").lower()
    if "ivory" in m or "bone" in m:        return 9
    if "ebony" in m:                       return 12
    if "brass" in m or "bronze" in m:      return 18
    if "cast metal" in m or "copper" in m: return 16
    if "music box" in t:                   return 20
    if "glass" in m:                       return 9
    if "porcelain" in m and ("large" in t or "set of" in t): return 28
    if "porcelain" in m or "ceramic" in m: return 13
    if "polyresin" in m or "resin" in m:   return 9
    if "celluloid" in m:                   return 4
    if "wood" in m:                        return 11
    if "print" in t or "painting" in t or "paper" in m: return 4
    if "soap" in t:                        return 11
    return 11


def box_size(oz: int, fragile: bool) -> str:
    wrap = "bubble-wrapped" if fragile else "padded"
    if oz < 13:   return f"6×6×4 box, {wrap}"
    if oz < 24:   return f"8×6×4 box, {wrap}"
    if oz < 48:   return f"10×8×6 box, double-{wrap}"
    return f"14×12×8 double-wall box, double-{wrap}"


def ship_svc(oz: int) -> str:
    if oz < 16:  return "USPS First Class Package"
    if oz < 80:  return "USPS Priority Mail"
    return "UPS Ground"


def fragile_flag(material: str) -> str:
    return "YES" if any(m in (material or "").lower()
                        for m in ["glass", "porcelain", "ceramic", "crystal"]) else "NO"


def packing_notes(item_type: str, fragile: str, is_lot: bool, n: int) -> str:
    base = "Double-box fragile — inner box cushioned with 2\" foam on all sides" if fragile == "YES" \
        else "Single box with bubble wrap and kraft fill"
    if is_lot:
        return f"{base}. Wrap each of the {n} pieces individually with dividers between pieces."
    return base


def derive_color(material: str, item_type: str) -> str:
    m, t = (material or "").lower(), (item_type or "").lower()
    for pair in [("gray","Gray"),("grey","Gray"),("gold","Gold"),("brass","Gold"),
                 ("ivory","Ivory/Cream"),("green","Green"),("seafoam","Seafoam"),
                 ("ebony","Black/Ebony"),("copper","Copper"),("bisque","Gray/Bisque"),
                 ("white","White/Ivory"),("dark","Dark Brown/Black")]:
        if pair[0] in m or pair[0] in t: return pair[1]
    return ""


def derive_production_tech(material: str, condition: str) -> str:
    m, c = (material or "").lower(), (condition or "").lower()
    parts = []
    if "hand-painted" in c or "hand painted" in c or "hand-painted" in m: parts.append("Hand-painted")
    if "hand carved" in c or "hand-carved" in m or "carved" in m:         parts.append("Hand-carved")
    if "hand blown" in m or "art glass" in m:                              parts.append("Hand-blown glass")
    if "cast" in m:                                                        parts.append("Cast")
    if "bisque" in m:                                                      parts.append("Bisque fired")
    if "celluloid" in m:                                                   parts.append("Celluloid molded")
    if "studio pottery" in m or "studio pottery" in c:                     parts.append("Studio pottery wheel/hand")
    return ", ".join(parts) if parts else "Machine-made / manufactured"


def derive_country(maker: str, marks: str) -> str:
    m, mk = (maker or "").lower(), (marks or "").lower()
    if "andrea by sadek" in m or "japan" in mk: return "Japan"
    if "lenox" in m or "aramis" in m:           return "USA"
    if "sandicast" in m:                        return "USA"
    if "reuge" in m:                            return "Switzerland"
    if "ubisoft" in m or "ubi workshop" in m:   return "Canada/USA"
    if "uruguay" in m or "hecho en uruguay" in mk: return "Uruguay"
    if "india" in m or "rajasthani" in m or "anglo-indian" in m: return "India"
    if "england" in mk or "english" in m:       return "England"
    if "italy" in mk:                           return "Italy"
    if "china" in mk:                           return "China"
    return ""


def etsy_era(era: str) -> str:
    e = (era or "").lower()
    for d in ["1920","1930","1940","1950","1960","1970","1980","1990"]:
        if d in e: return f"{d}s"
    if any(x in e for x in ["victorian","1800","edwardian"]): return "Before 1900s"
    if "1910" in e: return "1910s"
    return "Vintage (20-99 years old)"


def etsy_section(item_type: str, keywords: list, material: str) -> str:
    t, m = (item_type or "").lower(), (material or "").lower()
    kw = " ".join(keywords).lower()
    if "glass" in t or "art glass" in kw:                  return "Glass Collection"
    if any(s in t+kw for s in ["folk art","hand-carved","handmade","clay","carved wood"]): return "Folk Art & Artisan"
    if "brass" in t or "metal" in t or "gold" in kw:       return "Brass & Metal"
    if "music box" in t:                                    return "Music Boxes"
    if "print" in t or "painting" in t:                     return "Art & Prints"
    if "porcelain" in t or "ceramic" in t:                  return "Porcelain & Ceramic"
    return "Vintage Collectibles"


DIFF_NOTES = {
    "eBay→Etsy":    "eBay listing: front-load brand name and model number for keyword search, "
                    "price at market median. Etsy listing: lead with visual story and styling context, "
                    "price 10–15% higher — Etsy buyers pay for curation.",
    "eBay→Ruby":    "eBay listing: use keyword-rich title with maker and era. "
                    "Ruby Lane listing: formal appraisal language, emphasize provenance and materials, "
                    "price 20–30% above eBay — Ruby Lane audience pays for authentication.",
    "eBay→Poshmark":"eBay: collectible framing, keyword-heavy title. "
                    "Poshmark: lead with brand and luxury angle, list under 'Grooming & Health', "
                    "price slightly higher and note sealed/new condition.",
    "Etsy→eBay":    "Etsy listing: narrative tone, era story, suggest styling. "
                    "eBay listing: strip narrative, front-load keywords and maker name, "
                    "price at or just below eBay market median.",
    "Ruby→eBay":    "Ruby Lane: formal description with provenance, materials, measurements. "
                    "eBay listing: keyword-heavy, note material and era prominently, "
                    "price lower than Ruby Lane to reflect platform audience.",
}


# ── Platform evaluator ────────────────────────────────────────────────────────

BRAND_NAMES   = ["andrea by sadek","sandicast","lenox","aramis","reuge","ubisoft","ubi workshop","cheri blum"]
FOLK_SIGNALS  = ["folk art","handmade","hand-carved","hand carved","artisan","studio pottery","handcrafted","clay"]
AESTHETIC_SIG = ["cottagecore","bohemian","boho","eclectic","global","world art","mid century","vintage decor"]

def evaluate_platforms(row: dict, is_lot: bool = False) -> dict:
    maker  = (row.get("probable_maker") or "").lower()
    price  = float(row.get("suggested_price") or row.get("lot_price") or 0)
    mat    = (row.get("material") or "").lower()
    era    = (row.get("probable_era") or "").lower()
    kw_raw = row.get("style_keywords") or "[]"
    kws    = json.loads(kw_raw) if isinstance(kw_raw, str) else (kw_raw or [])
    kw_str = " ".join(kws).lower()
    itype  = (row.get("item_type") or row.get("lot_name") or "").lower()
    conf   = row.get("confidence_score") or 100

    has_fine   = any(x in mat or x in itype for x in ["ivory","ebony","sterling","bone inlay"])
    is_pre1950 = any(d in era for d in ["1800s","1900s","1910s","1920s","1930s","1940s","victorian","edwardian","anglo-indian"])
    is_branded = any(b in maker for b in BRAND_NAMES)
    is_gaming  = any(g in maker for g in ["ubisoft","ubi workshop"]) or "usb" in itype
    is_folk    = any(s in itype or s in kw_str for s in FOLK_SIGNALS)
    is_aesth   = any(s in kw_str for s in AESTHETIC_SIG)
    is_aramis  = "aramis" in maker

    primary   = "eBay"
    secondary = []
    specialty = []
    notes     = []

    # Specialty: fine materials
    if has_fine and price >= 25:
        primary = "Ruby Lane"
        specialty.append("1stDibs")
        secondary.append("eBay")
        notes.append("Fine materials (ivory/ebony/bone) qualify for specialty platforms")
        if "ivory" in mat:
            notes.append("⚠ eBay restricts ivory — confirm pre-1947 provenance before eBay listing")
    elif is_pre1950 and price >= 60:
        primary = "Ruby Lane"
        secondary.append("eBay")
        notes.append("Pre-1950 antique at qualifying price for Ruby Lane")

    # Poshmark: luxury grooming
    if is_aramis:
        (specialty if primary != "eBay" else secondary).append("Poshmark")
        notes.append("Aramis luxury brand suits Poshmark 'Grooming & Health'")

    # Etsy vs eBay (only if not already specialty)
    if primary not in ("Ruby Lane","1stDibs"):
        if is_folk and not is_branded:
            primary = "Etsy"; secondary.append("eBay")
            notes.append("Handmade/folk piece — story-driven Etsy audience is best fit")
        elif is_aesth and not is_branded and not is_gaming:
            primary = "Etsy"; secondary.append("eBay")
            notes.append("Strong decorative/aesthetic appeal suits Etsy curated-shelf buyer")
        elif is_branded or is_gaming:
            primary = "eBay"
            if is_aesth: secondary.append("Etsy")
            notes.append("Brand name drives keyword discovery on eBay")
        else:
            primary = "eBay"
            if is_aesth or is_folk: secondary.append("Etsy")

    # FB Marketplace override
    if row.get("yard_sale"):
        primary = "Facebook Marketplace"; secondary = []
        notes.append("Flagged for local/yard sale — shipping cost exceeds value")
    elif price < 12 and primary == "eBay" and not is_branded:
        primary = "Facebook Marketplace"
        notes.append("Low value relative to estimated shipping cost")

    # Lots: respect user-assigned platform, add cross-list if sensible
    if is_lot:
        assigned = row.get("suggested_platform") or primary
        primary  = assigned
        if assigned == "Etsy" and price >= 35 and "eBay" not in secondary:
            secondary.append("eBay")
        elif assigned == "eBay" and (is_folk or is_aesth) and "Etsy" not in secondary:
            secondary.append("Etsy")

    secondary = [s for s in dict.fromkeys(secondary) if s != primary]
    specialty  = [s for s in dict.fromkeys(specialty)  if s != primary]

    return {
        "primary":      primary,
        "secondary":    secondary,
        "specialty":    specialty,
        "is_cross_list":bool(secondary or specialty),
        "notes":        " | ".join(notes),
    }


# ── Row builders ──────────────────────────────────────────────────────────────

def lot_item_summaries(item_ids: list[str]) -> list[str]:
    summaries = []
    for iid in item_ids:
        item = db.get_item(iid)
        if not item: continue
        t = item.get("item_type") or iid
        e = item.get("probable_era") or ""
        m = item.get("material") or ""
        summaries.append(f"{t}{' (' + e + ')' if e else ''}{' — ' + m if m else ''}")
    return summaries


def enrich_lot(lot: dict) -> dict:
    """Add synthesized fields to a lot row from its component items."""
    item_ids = json.loads(lot.get("item_ids") or "[]")
    items = [db.get_item(iid) for iid in item_ids if db.get_item(iid)]
    if not items: return lot

    mats = list(dict.fromkeys(
        m for i in items for m in [(i.get("material") or "").split(",")[0].strip()] if m))
    all_kws = []
    for i in items:
        all_kws.extend(json.loads(i.get("style_keywords") or "[]"))
    eras = [i.get("probable_era") or "" for i in items if i.get("probable_era")]
    avg_conf = int(sum(i.get("confidence_score") or 80 for i in items) / len(items))
    comp_lows    = [i.get("comp_low") or 0    for i in items]
    comp_meds    = [i.get("comp_median") or 0 for i in items]
    comp_highs   = [i.get("comp_high") or 0   for i in items]

    enriched = dict(lot)
    enriched.update({
        "material":          ", ".join(mats[:3]),
        "style_keywords":    json.dumps(list(dict.fromkeys(all_kws))[:8]),
        "probable_era":      eras[0] if eras else "",
        "confidence_score":  avg_conf,
        "comp_low":          round(sum(comp_lows), 2),
        "comp_median":       round(sum(comp_meds), 2),
        "comp_high":         round(sum(comp_highs), 2),
        "comp_sample_size":  min(i.get("comp_sample_size") or 0 for i in items),
        "makers_marks_observed": "Varies — see individual piece descriptions",
        "probable_maker":    ", ".join(dict.fromkeys(
            i.get("probable_maker") or "" for i in items if i.get("probable_maker"))),
        "condition_notes":   "All pieces in good used/vintage condition. "
                             "See lot description for individual condition notes.",
        "item_type":         lot.get("lot_name") or "Lot",
        "_item_summaries":   lot_item_summaries(item_ids),
        "_n_items":          len(items),
        "_item_ids_list":    item_ids,
    })
    return enriched


def build_ebay_row(row: dict, is_lot: bool, plat: dict, row_id: int) -> dict:
    kw_raw  = row.get("style_keywords") or "[]"
    kws     = json.loads(kw_raw) if isinstance(kw_raw, str) else (kw_raw or [])
    mat     = row.get("material") or ""
    itype   = row.get("item_type") or row.get("lot_name") or ""
    maker   = row.get("probable_maker") or ""
    marks   = row.get("makers_marks_observed") or ""
    era     = row.get("probable_era") or ""
    cond    = row.get("condition_notes") or ""
    price   = float(row.get("suggested_price") or row.get("lot_price") or 0)
    n       = row.get("_n_items") or 1
    photos  = get_photos(row, is_lot)

    raw_title = (row.get("ebay_title") or row.get("lot_name") or "")
    if is_lot and not raw_title.lower().startswith("lot") and "collection" not in raw_title.lower() and "bundle" not in raw_title.lower():
        raw_title = f"Lot of {n}: {raw_title}"

    woz   = weight_oz(mat, itype) * (n if is_lot else 1)
    frag  = fragile_flag(mat)
    svc   = ship_svc(woz)
    cond_label = "New" if any(x in cond.lower() for x in ["sealed","never used","factory-fresh","new"]) else "Used"

    summaries = row.get("_item_summaries") or []
    html_desc = fmt_html(row.get("description") or row.get("lot_description") or "", summaries if is_lot else None)

    d = {**photo_dict(photos)}
    d.update({
        "row_id":          row_id,
        "ebay_title":      trunc(raw_title, 80),
        "subtitle":        trunc(f"{era} · {', '.join(kws[:2])}" if kws else era, 55),
        "html_description":html_desc,
        "price":           f"{price:.2f}",
        "format":          "FixedPrice",
        "duration":        "GTC",
        "quantity":        "1",
        "condition":       cond_label,
        "condition_desc":  cond[:500] if cond else "",
        "upc":             "Does not apply",
        "material":        mat[:100],
        "subject":         "Elephant" if "elephant" in itype.lower() or any("elephant" in k.lower() for k in kws) else itype.split()[0] if itype else "",
        "theme":           ", ".join(kws[:3]),
        "color":           derive_color(mat, itype),
        "origin":          derive_country(maker, marks),
        "item_type_short": itype[:60],
        "production_style":("Folk Art" if any(s in itype.lower() for s in ["folk","handmade","carved"]) else
                            "Studio Art" if "studio" in itype.lower() else
                            "Commercial / Mass Produced"),
        "production_tech": derive_production_tech(mat, cond),
        "marks":           marks[:100] if marks else "None visible",
        "country_origin":  derive_country(maker, marks),
        "features":        ", ".join(kws[2:5]) if len(kws) > 2 else "",
        "time_period":     era,
        "mpn":             "Does not apply",
        "pattern":         "",
        "product_line":    maker[:60] if maker else "",
        "year_mfr":        re.search(r"\d{4}", era).group() if re.search(r"\d{4}", era) else "",
        "style":           ", ".join(kws[:2]),
        "finish":          next((w for w in ["matte","glossy","bisque","antiqued","gold-tone","natural","painted"]
                                 if w in (mat+cond).lower()), ""),
        "weight_oz":       woz,
        "box_size":        box_size(woz, frag=="YES"),
        "ship_svc":        svc,
        "fragile":         frag,
        "pack_notes":      packing_notes(itype, frag, is_lot, n),
        "ship_type":       "Calculated",
        "handling":        "3 business days",
        "location":        "Saint Charles, MO",
        "photo_notes":     ("First photo = composite grid thumbnail. Photos 2-5 = individual piece detail shots." if is_lot else
                            "Lead with clearest, best-lit angle. Mark photo first."),
        "confidence":      row.get("confidence_score") or "",
        "comp_low":        f"${row.get('comp_low') or 0:.2f}" if row.get('comp_low') else "",
        "comp_median":     f"${row.get('comp_median') or 0:.2f}" if row.get('comp_median') else "",
        "comp_high":       f"${row.get('comp_high') or 0:.2f}" if row.get('comp_high') else "",
        "comp_n":          row.get("comp_sample_size") or "",
        "secondary":       ", ".join(plat["secondary"] + plat["specialty"]),
        "lot_id":          row.get("lot_id") or (row.get("lot_id") if is_lot else ""),
        "special_notes":   plat["notes"],
    })
    return d


def build_etsy_row(row: dict, is_lot: bool, plat: dict, row_id: int) -> dict:
    kw_raw = row.get("style_keywords") or "[]"
    kws    = json.loads(kw_raw) if isinstance(kw_raw, str) else (kw_raw or [])
    mat    = row.get("material") or ""
    itype  = row.get("item_type") or row.get("lot_name") or ""
    maker  = row.get("probable_maker") or ""
    era    = row.get("probable_era") or ""
    price  = float(row.get("suggested_price") or row.get("lot_price") or 0)
    photos = get_photos(row, is_lot)

    # Etsy tags: up to 13, combine style keywords + descriptive terms
    existing_tags = json.loads(row.get("tags") or "[]") if not is_lot else []
    tag_pool      = list(dict.fromkeys(existing_tags + kws))[:13]

    # Etsy description: plain text, keep the narrative description
    etsy_desc = row.get("description") or row.get("lot_description") or ""
    if is_lot and row.get("_item_summaries"):
        summaries = row["_item_summaries"]
        etsy_desc += "\n\nThis lot includes:\n" + "\n".join(f"• {s}" for s in summaries)
        etsy_desc += "\n\nSold as a set only — will not separate."

    raw_title = row.get("etsy_title") or row.get("lot_name") or ""
    etsy_title = trunc(raw_title, 140)

    is_brand = any(b in (maker or "").lower() for b in BRAND_NAMES)
    who_made = "A company or brand" if is_brand else "Unknown craftsperson"

    d = {**photo_dict(photos)}
    d.update({
        "row_id":          row_id,
        "etsy_title":      etsy_title,
        "etsy_description":etsy_desc,
        "price":           f"{price:.2f}",
        "tags":            ", ".join(tag_pool),
        "materials":       mat[:200],
        "etsy_category":   ("Home & Living > Home Decor > Figurines" if "figurine" in itype.lower() else
                            "Art & Collectibles > Collectibles > Figurines" if "figurine" in itype.lower() else
                            "Art & Collectibles > Art > Paintings" if "painting" in itype.lower() else
                            "Home & Living > Home Decor > Decorative Objects"),
        "etsy_era":        etsy_era(era),
        "who_made_it":     who_made,
        "quantity":        "1",
        "renewal":         "Automatic",
        "ship_profile":    f"Calculated shipping — est. {weight_oz(mat,itype) * (row.get('_n_items') or 1)} oz packed",
        "etsy_section":    etsy_section(itype, kws, mat),
        "photo_notes":     ("Etsy thumbnail (Photo 1) = most visually striking angle, not necessarily front-facing. "
                            "Use natural light. Composite grid as first photo for lots." if is_lot else
                            "Etsy thumbnail (Photo 1) should be the most visually striking angle, "
                            "ideally at eye level with soft natural light."),
        "secondary":       ", ".join(plat["secondary"] + plat["specialty"]),
        "special_notes":   plat["notes"],
    })
    return d


def build_fb_row(row: dict, is_lot: bool, plat: dict, row_id: int) -> dict:
    price  = float(row.get("suggested_price") or row.get("lot_price") or 0)
    itype  = row.get("item_type") or row.get("lot_name") or ""
    cond   = row.get("condition_notes") or ""
    photos = get_photos(row, is_lot)

    fb_title = trunc(row.get("fb_marketplace_title") or row.get("lot_name") or itype, 80)
    # Conversational FB description
    short_cond = cond.split(".")[0] if cond else "good condition"
    fb_desc = (
        f"{fb_title}, {short_cond.lower()}. "
        f"Pickup in St. Charles, MO. "
        f"{'Message with questions!' if price < 30 else 'Happy to answer questions before purchase.'}"
    )
    if is_lot:
        fb_desc += f" Lot of {row.get('_n_items') or 'several'} pieces, selling as a set."

    fb_cat = ("Antiques & Collectibles" if any(x in itype.lower() for x in ["figurine","elephant","vintage"]) else
              "Home Goods" if any(x in itype.lower() for x in ["decor","print","painting","dish"]) else
              "Antiques & Collectibles")

    d = {**photo_dict(photos)}
    d.update({
        "row_id":       row_id,
        "fb_title":     fb_title,
        "fb_description": fb_desc,
        "fb_price":     str(round(price * 0.9 / 5) * 5 or 5),
        "fb_category":  fb_cat,
        "condition":    "Good" if "chip" not in (cond or "").lower() else "Fair",
        "location":     "Saint Charles, MO",
        "neg_floor":    f"${price * 0.7:.0f}",
        "special_notes":plat["notes"],
    })
    return d


def build_spec_row(row: dict, is_lot: bool, plat: dict, platform: str, row_id: int) -> dict:
    price   = float(row.get("suggested_price") or row.get("lot_price") or 0)
    mat     = row.get("material") or ""
    itype   = row.get("item_type") or row.get("lot_name") or ""
    maker   = row.get("probable_maker") or ""
    era     = row.get("probable_era") or ""
    marks   = row.get("makers_marks_observed") or ""
    photos  = get_photos(row, is_lot)

    # Specialty prices: Ruby Lane/1stDibs = 30% above eBay; Poshmark = eBay price
    spec_price = price * 1.3 if platform in ("Ruby Lane","1stDibs") else price

    # Formal description for Ruby Lane / 1stDibs
    raw_desc = row.get("description") or row.get("lot_description") or ""
    if platform in ("Ruby Lane","1stDibs"):
        spec_desc = (
            f"MAKER/ORIGIN: {maker or 'Unknown maker'}. "
            f"ERA: {era}. "
            f"MATERIAL: {mat}. "
            f"MARKS: {marks or 'None visible'}.\n\n"
        ) + raw_desc
    else:
        spec_desc = raw_desc

    key_pts = []
    if any(x in mat.lower() for x in ["ivory","ebony","bone","sterling"]): key_pts.append("Fine natural materials")
    if any(x in (era or "").lower() for x in ["1800s","1900s","1910s","1920s","1930s","1940s"]): key_pts.append("Pre-1950 antique")
    if marks and marks.lower() not in ["none visible","n/a"]: key_pts.append(f"Maker's marks present: {marks[:40]}")
    if "carved" in itype.lower(): key_pts.append("Hand-carved craftsmanship")
    if "set of" in itype.lower() or "graduated" in itype.lower(): key_pts.append("Complete graduated/matched set")

    reqs = {
        "Ruby Lane":  "Requires title, full description with measurements, condition grade (Excellent/Very Good/Good/Fair), "
                      "at least 4 photos including marks/backstamp. No reserve pricing. Seller account required ($25/mo).",
        "1stDibs":    "Requires dealer account. Strict image standards (white background, min 1000px). "
                      "Must include provenance statement and condition report.",
        "Poshmark":   "List under 'Grooming & Health > Fragrance' or 'Other'. "
                      "Lead with 'SEALED' or 'NWT' in title. Poshmark takes 20% on sales > $15.",
    }.get(platform, "")

    d = {**photo_dict(photos)}
    d.update({
        "row_id":           row_id,
        "spec_platform":    platform,
        "spec_title":       trunc(row.get("ebay_title") or row.get("lot_name") or itype, 80),
        "spec_description": spec_desc,
        "spec_price":       f"{spec_price:.2f}",
        "spec_category":    itype[:60],
        "key_points":       "\n".join(f"• {p}" for p in key_pts) if key_pts else "",
        "spec_requirements":reqs,
    })
    return d


def build_review_row(item: dict, issue: str, row_id: int) -> dict:
    photos = get_photos(item, False)
    return {
        "row_id":       row_id,
        "item_id":      item.get("item_id") or "",
        "item_type":    item.get("item_type") or "—",
        "maker":        item.get("probable_maker") or "—",
        "confidence":   item.get("confidence_score") or "—",
        "issue":        issue,
        "price":        f"${item.get('suggested_price') or 0:.2f}",
        "photo_1":      photos[0] if photos else "",
        "special_notes":item.get("condition_notes") or "",
    }


# ── Sheet writer ──────────────────────────────────────────────────────────────

def write_sheet(ws, col_defs: list, rows: list[dict],
                hdr_color: str, sec_colors: dict | None = None,
                flag_field: str = "confidence", flag_thresh: int = 70,
                note_field: str = "special_notes"):
    """Generic sheet writer with frozen header, auto-filter, section coloring."""
    hdr_fill   = _fill(hdr_color)
    hdr_font   = _font(bold=True, color=C.WHITE, size=10)
    center_al  = _align(h="center", v="center", wrap=False)
    wrap_al    = _align(wrap=True, v="top")

    # Write header row
    for ci, col_def in enumerate(col_defs, 1):
        label = col_def[0]
        sec   = col_def[2] if len(col_def) > 2 and isinstance(col_def[2], str) and len(col_def[2]) == 1 else None
        cell  = ws.cell(row=1, column=ci, value=label)
        cell.font      = hdr_font
        cell.alignment = center_al
        cell.fill      = _fill(sec_colors[sec]) if sec_colors and sec and sec in sec_colors else hdr_fill

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Write data rows
    for ri, row_data in enumerate(rows, 2):
        conf  = row_data.get(flag_field)
        notes = row_data.get(note_field) or ""
        is_low_conf = isinstance(conf, int) and conf < flag_thresh
        is_flagged  = bool(notes.strip())
        row_fill    = (_fill(C.FLAG_ROW) if is_flagged and is_low_conf else
                       _fill(C.LOW_CONF) if is_low_conf else
                       _fill(C.FLAG_ROW) if is_flagged else None)

        ws.row_dimensions[ri].height = 60

        for ci, col_def in enumerate(col_defs, 1):
            field = col_def[1]
            val   = row_data.get(field, "")
            cell  = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = wrap_al
            if is_low_conf or is_flagged:
                cell.fill = row_fill
                if is_low_conf or is_flagged:
                    cell.font = _font(bold=True) if (is_low_conf or is_flagged) else _font()

    # Column widths
    for ci, col_def in enumerate(col_defs, 1):
        width = col_def[-1] if isinstance(col_def[-1], (int, float)) else 18
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Section header label row above main header (for eBay only, where sec_colors apply)
    if sec_colors:
        sec_spans: dict[str, list[int]] = {}
        for ci, col_def in enumerate(col_defs, 1):
            sec = col_def[2] if len(col_def) > 2 and isinstance(col_def[2], str) and len(col_def[2]) == 1 else None
            if sec:
                sec_spans.setdefault(sec, []).append(ci)

        SEC_LABELS = {"A":"Section A — Listing Basics","B":"Section B — Item Specifics",
                      "C":"Section C — Shipping","D":"Section D — Photos","E":"Section E — Internal Notes"}
        ws.insert_rows(1)
        for sec, cols in sec_spans.items():
            start, end = min(cols), max(cols)
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(row=1, column=start, value=SEC_LABELS.get(sec, sec))
            cell.fill      = _fill(sec_colors[sec])
            cell.font      = _font(bold=True, size=10)
            cell.alignment = _align(h="center", wrap=False)
        ws.row_dimensions[1].height = 16
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(col_defs))}2"


# ── Main export ───────────────────────────────────────────────────────────────

def run_export() -> Path:
    db.init_db()
    all_items = db.get_all_items()
    all_lots  = db.get_all_lots()

    # Evaluate platforms for everything
    assessed_items = []
    for item in all_items:
        plat = evaluate_platforms(item, is_lot=False)
        assessed_items.append((item, plat, False))

    assessed_lots = []
    for lot in all_lots:
        enriched = enrich_lot(lot)
        plat = evaluate_platforms(enriched, is_lot=True)
        assessed_lots.append((enriched, plat, True))

    all_assessed = assessed_items + assessed_lots

    # Partition by destination sheet
    ebay_rows, etsy_rows, fb_rows = [], [], []
    spec_rows, review_rows = [], []
    cross_candidates: list[tuple] = []   # (name, primary, secondary, specialty, notes)

    # Row counters per sheet (for cross-ref)
    ebay_ref: dict[str, int] = {}
    etsy_ref: dict[str, int] = {}
    fb_ref:   dict[str, int] = {}
    spec_ref: dict[str, int] = {}

    # Needs review: unprocessed / flagged
    for item in all_items:
        issues = []
        if not item.get("ebay_title"):  issues.append("No listing title — not yet processed through pipeline")
        if not item.get("suggested_price") or item.get("suggested_price") == 0:
            issues.append("Price is $0 — needs identification/comp run")
        if item.get("needs_review"):    issues.append(f"Low confidence ({item.get('confidence_score')}%)")
        if issues:
            review_rows.append(build_review_row(item, "; ".join(issues), len(review_rows)+1))

    # Main routing
    for row, plat, is_lot in all_assessed:
        name = row.get("item_id") or row.get("lot_id") or "?"
        # Skip unprocessed (no title, no price)
        if not row.get("ebay_title") and not row.get("lot_name"):
            continue
        if not (row.get("suggested_price") or row.get("lot_price")):
            continue

        p = plat["primary"]

        if p == "eBay":
            row_n = len(ebay_rows) + 2
            ebay_rows.append(build_ebay_row(row, is_lot, plat, len(ebay_rows)+1))
            ebay_ref[name] = row_n
        elif p == "Etsy":
            row_n = len(etsy_rows) + 2
            etsy_rows.append(build_etsy_row(row, is_lot, plat, len(etsy_rows)+1))
            etsy_ref[name] = row_n
        elif p == "Facebook Marketplace":
            row_n = len(fb_rows) + 2
            fb_rows.append(build_fb_row(row, is_lot, plat, len(fb_rows)+1))
            fb_ref[name] = row_n
        elif p in ("Ruby Lane","1stDibs","Poshmark"):
            row_n = len(spec_rows) + 2
            spec_rows.append(build_spec_row(row, is_lot, plat, p, len(spec_rows)+1))
            spec_ref[name] = row_n

        # Secondary platforms also get a row on their sheet
        for sec in plat["secondary"]:
            if sec == "eBay":
                ebay_rows.append(build_ebay_row(row, is_lot, plat, len(ebay_rows)+1))
                ebay_ref[f"{name}_x"] = len(ebay_rows) + 1
            elif sec == "Etsy":
                etsy_rows.append(build_etsy_row(row, is_lot, plat, len(etsy_rows)+1))
                etsy_ref[f"{name}_x"] = len(etsy_rows) + 1
        for spec_p in plat["specialty"]:
            spec_rows.append(build_spec_row(row, is_lot, plat, spec_p, len(spec_rows)+1))
            spec_ref[f"{name}_{spec_p}"] = len(spec_rows) + 1

        # Track cross-list candidates
        if plat["is_cross_list"]:
            cross_candidates.append((row, plat, is_lot, name))

    # Cross-list sheet
    cross_rows = []
    for row, plat, is_lot, name in cross_candidates:
        label = row.get("ebay_title") or row.get("lot_name") or name
        all_secs  = plat["secondary"] + plat["specialty"]
        platforms = [plat["primary"]] + all_secs
        ref_maps  = {"eBay": ebay_ref, "Etsy": etsy_ref,
                     "Facebook Marketplace": fb_ref,
                     "Ruby Lane": spec_ref, "1stDibs": spec_ref, "Poshmark": spec_ref}

        # Build differentiation note
        key = f"{plat['primary']}→{'→'.join(all_secs[:1])}" if all_secs else ""
        diff = DIFF_NOTES.get(key, plat["notes"] or "Tailor title keywords to each platform's search behaviour.")

        cr = {"name": trunc(label, 60)}
        for i, p in enumerate(platforms[:3], 1):
            ref_key = name if i == 1 else f"{name}_x"
            sheet_row = ref_maps.get(p, {}).get(ref_key, ref_maps.get(p, {}).get(name, ""))
            cr[f"platform_{i}"] = p
            cr[f"row_p{i}"]     = f"Row {sheet_row}" if sheet_row else "—"
        for i in range(len(platforms)+1, 4):
            cr[f"platform_{i}"] = ""
            cr[f"row_p{i}"]     = ""
        cr["diff_notes"] = diff
        cross_rows.append(cr)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    ws_ebay   = wb.active;            ws_ebay.title   = "eBay Listings"
    ws_etsy   = wb.create_sheet("Etsy Listings")
    ws_fb     = wb.create_sheet("FB Marketplace")
    ws_spec   = wb.create_sheet("Specialty Platforms")
    ws_cross  = wb.create_sheet("Cross-List Candidates")
    ws_review = wb.create_sheet("Needs Review")

    write_sheet(ws_ebay,   EBAY_COLS,    ebay_rows,   C.EBAY_HDR,  SEC_COLORS,  "confidence", 70)
    write_sheet(ws_etsy,   ETSY_COLS,    etsy_rows,   C.ETSY_HDR,  None,        "confidence", 70)
    write_sheet(ws_fb,     FB_COLS,      fb_rows,     C.FB_HDR,    None,        "confidence", 70)
    write_sheet(ws_spec,   SPEC_COLS,    spec_rows,   C.SPEC_HDR,  None,        "confidence", 70)
    write_sheet(ws_cross,  CROSS_COLS,   cross_rows,  C.CROSS_HDR, None,        "confidence", 70)
    write_sheet(ws_review, REVIEW_COLS,  review_rows, C.REVIEW_HDR,None,        "confidence", 70)

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)
    return wb, ebay_rows, etsy_rows, fb_rows, spec_rows, cross_rows, review_rows


def main():
    print("\n=== Master Listings Export ===\n")
    db.init_db()
    wb, ebay_r, etsy_r, fb_r, spec_r, cross_r, review_r = run_export()

    print(f"  eBay Listings          : {len(ebay_r):3d} rows")
    print(f"  Etsy Listings          : {len(etsy_r):3d} rows")
    print(f"  FB Marketplace         : {len(fb_r):3d} rows")
    print(f"  Specialty Platforms    : {len(spec_r):3d} rows")
    print(f"  Cross-List Candidates  : {len(cross_r):3d} rows")
    print(f"  Needs Review           : {len(review_r):3d} rows")
    print(f"\nSaved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
